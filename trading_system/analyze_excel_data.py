"""
Analyze Historical Trading Data from Excel
Loads data from TRADING_DATA_REPORT.xlsx and runs all indicators to generate signals
"""

import sys
import io

# Fix Windows console encoding
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from utils import connect_to_database
from master_confluence import analyze_master_confluence

# Configuration
INPUT_FILE = Path('../doc/TRADING_DATA_REPORT.xlsx')
OUTPUT_FILE = Path('../doc/SIGNALS_FROM_HISTORICAL_DATA.xlsx')

def load_excel_data():
    """
    Load all data from Excel file

    Returns:
        dict: {symbol: DataFrame with OHLCV data}
    """
    print(f"\n📥 Loading data from {INPUT_FILE}...")
    print("=" * 80)

    if not INPUT_FILE.exists():
        print(f"❌ File not found: {INPUT_FILE}")
        return None

    # Load all sheets except Summary
    wb = load_workbook(INPUT_FILE)
    symbols = [sheet for sheet in wb.sheetnames if sheet != 'Summary']

    print(f"Found {len(symbols)} symbol sheets: {', '.join(symbols)}")

    data_by_symbol = {}

    for symbol in symbols:
        print(f"\n  Loading {symbol}...", end=' ')
        df = pd.read_excel(INPUT_FILE, sheet_name=symbol)

        # Ensure we have required columns
        required_cols = ['symbol', 'timeframe', 'timestamp', 'open', 'high', 'low', 'close', 'volume']
        if not all(col in df.columns for col in required_cols):
            print(f"❌ Missing required columns")
            continue

        # Convert datetime column to timestamp if needed
        if 'date' in df.columns:
            # Some timestamps are already Unix timestamps, some might be datetime objects
            df['timestamp'] = pd.to_datetime(df['date']).astype(int) / 10**9

        print(f"✅ {len(df)} rows")
        data_by_symbol[symbol] = df

    wb.close()

    print("\n" + "=" * 80)
    print(f"✅ Loaded {len(data_by_symbol)} symbols with {sum(len(df) for df in data_by_symbol.values())} total rows")

    return data_by_symbol

def clear_and_load_data_to_db(data_by_symbol):
    """
    Clear existing data and load Excel data into database

    Args:
        data_by_symbol: Dictionary of {symbol: DataFrame}
    """
    print("\n💾 Loading data into database...")
    print("=" * 80)

    conn = connect_to_database()
    cursor = conn.cursor()

    # Get list of symbols
    symbols = list(data_by_symbol.keys())

    # Clear existing data for these symbols
    print(f"\n  Clearing existing data for: {', '.join(symbols)}")
    placeholders = ','.join(['?' for _ in symbols])
    cursor.execute(f"DELETE FROM price_data WHERE symbol IN ({placeholders})", symbols)
    deleted = cursor.rowcount
    print(f"  ✅ Cleared {deleted} old rows")

    # Insert new data
    total_inserted = 0

    for symbol, df in data_by_symbol.items():
        print(f"\n  Inserting {symbol}...", end=' ')

        for _, row in df.iterrows():
            cursor.execute("""
                INSERT INTO price_data
                (symbol, timeframe, timestamp, open, high, low, close, volume, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            """, (
                row['symbol'],
                row['timeframe'],
                float(row['timestamp']),
                float(row['open']),
                float(row['high']),
                float(row['low']),
                float(row['close']),
                float(row['volume'])
            ))
            total_inserted += 1

        print(f"✅ {len(df)} rows")

    conn.commit()
    conn.close()

    print("\n" + "=" * 80)
    print(f"✅ Inserted {total_inserted:,} total rows into database")

def generate_signals_from_db():
    """
    Generate signals for all symbol/timeframe combinations in database

    Returns:
        list: List of signal dictionaries
    """
    print("\n🔍 Generating signals from loaded data...")
    print("=" * 80)

    conn = connect_to_database()
    cursor = conn.cursor()

    # Get all symbol/timeframe combinations
    cursor.execute("""
        SELECT DISTINCT symbol, timeframe
        FROM price_data
        ORDER BY symbol, timeframe
    """)

    combinations = cursor.fetchall()
    all_signals = []

    print(f"\nFound {len(combinations)} symbol/timeframe combinations to analyze\n")

    for idx, (symbol, timeframe) in enumerate(combinations, 1):
        print(f"[{idx}/{len(combinations)}] Analyzing {symbol} ({timeframe})...", end=' ')

        # Analyze this symbol/timeframe with master confluence
        result = analyze_master_confluence(conn, symbol, timeframe)

        if result and result.get('confluence'):
            confluence = result['confluence']
            score = confluence.get('score', 0)

            # Include signals with score > 0
            if score > 0:
                # Get all individual signals for detailed breakdown
                all_indicator_signals = []

                # Hull signals
                for sig in result.get('hull_signals', []):
                    all_indicator_signals.append(f"Hull: {sig['description']}")

                # AO signals
                for sig in result.get('ao_signals', []):
                    all_indicator_signals.append(f"AO: {sig['description']}")

                # Alligator signals
                for sig in result.get('alligator_signals', []):
                    all_indicator_signals.append(f"Alligator: {sig['description']}")

                # Ichimoku signals
                for sig in result.get('ichimoku_signals', []):
                    all_indicator_signals.append(f"Ichimoku: {sig['description']}")

                # Volume info
                volume_info = "NORMAL"
                volume_ratio = 0
                if result.get('volume_signals'):
                    latest_vol = result['volume_signals'][-1]
                    volume_info = latest_vol['level']
                    volume_ratio = latest_vol['ratio']

                signal_data = {
                    'Symbol': symbol,
                    'Timeframe': timeframe,
                    'Timestamp': result.get('timestamp', 0),
                    'Date': datetime.fromtimestamp(result.get('timestamp', 0)).strftime('%Y-%m-%d %H:%M:%S'),
                    'Price': result.get('price', 0),
                    'Direction': '🌪️ Bullish' if confluence.get('primary_system') == 'wind_catcher' else '🌊 Bearish',
                    'Score': round(score, 2),
                    'Classification': confluence.get('classification', 'UNKNOWN'),
                    'Signal_Count': confluence.get('signal_count', 0),
                    'Hull_Signals': len(result.get('hull_signals', [])),
                    'AO_Signals': len(result.get('ao_signals', [])),
                    'Alligator_Signals': len(result.get('alligator_signals', [])),
                    'Ichimoku_Signals': len(result.get('ichimoku_signals', [])),
                    'Volume_Level': volume_info,
                    'Volume_Ratio': round(volume_ratio, 2),
                    'All_Signals': ' | '.join(all_indicator_signals[:10])  # First 10 signals
                }

                all_signals.append(signal_data)

                emoji = confluence.get('emoji', '💫')
                print(f"{emoji} {confluence.get('classification')} (Score: {score:.1f})")
            else:
                print("⚪ No signals")
        else:
            print("❌ Insufficient data")

    conn.close()

    print("\n" + "=" * 80)
    print(f"✅ Generated {len(all_signals)} signals")

    return all_signals

def export_signals_to_excel(signals):
    """
    Export signals to Excel with multiple sheets for easy analysis

    Args:
        signals: List of signal dictionaries
    """
    print(f"\n📊 Exporting {len(signals)} signals to Excel...")
    print("=" * 80)

    if not signals:
        print("⚠️  No signals to export!")
        return

    # Create DataFrame
    df = pd.DataFrame(signals)

    # Sort by score (highest first), then by symbol and date
    df = df.sort_values(['Score', 'Symbol', 'Timestamp'], ascending=[False, True, False])

    # Ensure output directory exists
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Sheet 1: All Signals (sorted by score)
        df.to_excel(writer, sheet_name='All Signals', index=False)

        # Sheet 2: High Quality Signals Only (Score >= 1.8)
        high_quality = df[df['Score'] >= 1.8].copy()
        if not high_quality.empty:
            high_quality.to_excel(writer, sheet_name='High Quality (≥1.8)', index=False)

        # Sheet 3: Summary Statistics
        summary_data = {
            'Metric': [
                'Total Signals Found',
                'PERFECT (≥3.0)',
                'EXCELLENT (≥2.5)',
                'VERY GOOD (≥1.8)',
                'GOOD (≥1.2)',
                'WEAK (<1.2)',
                '',
                'Bullish Signals',
                'Bearish Signals',
                '',
                'Symbols Analyzed',
                'Timeframes',
                'Date Range',
                'Generated On'
            ],
            'Value': [
                len(df),
                len(df[df['Score'] >= 3.0]),
                len(df[df['Score'] >= 2.5]),
                len(df[df['Score'] >= 1.8]),
                len(df[df['Score'] >= 1.2]),
                len(df[df['Score'] < 1.2]),
                '',
                len(df[df['Direction'].str.contains('Bullish')]),
                len(df[df['Direction'].str.contains('Bearish')]),
                '',
                ', '.join(sorted(df['Symbol'].unique())),
                ', '.join(sorted(df['Timeframe'].unique())),
                f"{df['Date'].min()} to {df['Date'].max()}",
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

        # Sheet 4-N: Per-symbol sheets (sorted by date)
        for symbol in sorted(df['Symbol'].unique()):
            symbol_df = df[df['Symbol'] == symbol].sort_values('Timestamp', ascending=False)
            # Sanitize sheet name (Excel doesn't allow / : * ? [ ] characters)
            safe_sheet_name = symbol.replace('/', '_').replace(':', '_')[:31]  # Max 31 chars
            symbol_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

    print(f"✅ Excel report saved to: {OUTPUT_FILE}")
    print(f"   📄 Main sheets: All Signals, High Quality, Summary")
    print(f"   📄 Per-symbol sheets: {', '.join(sorted(df['Symbol'].unique()))}")

    # Print quick summary
    print("\n" + "=" * 80)
    print("📈 SIGNAL SUMMARY:")
    print(f"   Total Signals: {len(df)}")
    print(f"   PERFECT (≥3.0): {len(df[df['Score'] >= 3.0])}")
    print(f"   EXCELLENT (≥2.5): {len(df[df['Score'] >= 2.5])}")
    print(f"   VERY GOOD (≥1.8): {len(df[df['Score'] >= 1.8])}")
    print(f"   GOOD (≥1.2): {len(df[df['Score'] >= 1.2])}")
    print("=" * 80)

def main():
    """Main execution"""
    print("🚀 Historical Trading Data Signal Analyzer")
    print("=" * 80)
    print(f"📋 Input:  {INPUT_FILE}")
    print(f"📋 Output: {OUTPUT_FILE}")
    print("=" * 80)

    # Step 1: Load Excel data
    print("\n🔄 STEP 1: Load Excel Data")
    data_by_symbol = load_excel_data()

    if not data_by_symbol:
        print("❌ Failed to load data from Excel")
        return

    # Step 2: Load data into database
    print("\n🔄 STEP 2: Load Data into Database")
    clear_and_load_data_to_db(data_by_symbol)

    # Step 3: Generate signals using master confluence system
    print("\n🔄 STEP 3: Generate Signals")
    signals = generate_signals_from_db()

    # Step 4: Export to Excel
    print("\n🔄 STEP 4: Export Results")
    export_signals_to_excel(signals)

    print("\n" + "=" * 80)
    print("✅ COMPLETE! Signal analysis finished successfully!")
    print(f"📊 Open the file: {OUTPUT_FILE}")
    print("\n💡 TIP: Use the 'Date' column to find signals in TradingView")
    print("        The 'All_Signals' column shows exactly which patterns were detected")
    print("=" * 80)

if __name__ == "__main__":
    main()
